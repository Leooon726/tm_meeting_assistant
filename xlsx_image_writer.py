import uuid
import os

from openpyxl.drawing.image import Image
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPoint2D, XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU, cm_to_EMU

from excel_utils import coordinate_string_to_index,get_merged_cell_size,is_cell_in_block,subtract_coordinates,add_coordinates

_CAHR_WIDTH_PIXELS = 7.1
_DEFAULT_PADDING_IN_PIXEL = 1

class XlsxBlockImageWriter:
    def __init__(self,template_sheet,source_block_position,target_block_start_coord,template_image_loader):
        self.template_sheet = template_sheet
        self.image_loader = template_image_loader
        self.source_block_start_coord = source_block_position['start_coord']
        self.source_block_end_coord = source_block_position['end_coord']
        self.target_block_start_coord = target_block_start_coord

    def write(self,target_sheet):
        # get image source coords from template sheet block.
        image_source_coords = self._get_image_source_coords()
        for image_source_coord in image_source_coords:
            image_coord_offset = self._get_image_coord_offset(image_source_coord)
            image_tartget_coord = self._get_image_target_coords(image_coord_offset)
            self._add_image_for_cell(target_sheet,image_source_coord,image_tartget_coord)
        return target_sheet

    def _get_image_source_coords(self):
        image_source_coords = []
        start_col, start_row = coordinate_string_to_index(self.source_block_start_coord)
        end_col, end_row = coordinate_string_to_index(self.source_block_end_coord)
        for row in self.template_sheet.iter_rows(min_row=start_row,
                                            min_col=start_col,
                                            max_row=end_row,
                                            max_col=end_col):
            for cell in row:
                if not self.image_loader.image_in(cell.coordinate):
                    continue
                image_source_coords.append(cell.coordinate)
        return image_source_coords

    def _get_image_coord_offset(self,image_source_coord):
        image_offset = subtract_coordinates(image_source_coord,self.source_block_start_coord)
        return image_offset

    def _get_image_target_coords(self,image_offset):
        target_coord = add_coordinates(self.target_block_start_coord,image_offset)
        return target_coord

    def _add_image_for_cell(self,target_sheet,
                  source_cell_coord,
                  target_cell_coord):
        if not self.image_loader.image_in(source_cell_coord):
            print(source_cell_coord,"does not contain an image")
            return

        cell_size_in_pixels = self._get_merged_cell_size_in_pixels(self.template_sheet,source_cell_coord)

        for image_dict in self.image_loader.get_images_from_cell(source_cell_coord):
            image_path = image_dict['image_path']
            col_offset_in_EMU = image_dict['col_offset']
            row_offset_in_EMU = image_dict['row_offset']
            self._add_single_image(target_sheet,image_path,cell_size_in_pixels,target_cell_coord,col_offset_in_EMU,row_offset_in_EMU)

        return target_sheet

    @staticmethod
    def _get_merged_cell_size_in_pixels(sheet,cell_coord):
        cell_width_in_characters,cell_height_in_points = get_merged_cell_size(sheet,cell_coord)
        cell_height_in_pixels = XlsxBlockImageWriter._points_to_pixels(cell_height_in_points)
        
        cell_font_size = sheet[cell_coord].font.size
        cell_width_in_pixels = XlsxBlockImageWriter._characters_to_pixels(cell_width_in_characters,cell_font_size)

        return cell_width_in_pixels,cell_height_in_pixels

    def _add_single_image(self,target_sheet,image_path,cell_size_in_pixels,target_cell_coord,col_offset_in_EMU,row_offset_in_EMU):
        image = Image(image_path)
        self._resize_image_base_on_cell_size(image,cell_size_in_pixels)

        self._add_image_with_offset_and_padding(target_sheet,image,target_cell_coord,col_offset_in_EMU,row_offset_in_EMU)

    @staticmethod
    def _add_image_with_offset_and_padding(sheet,image,cell_coord,col_offset_in_EMU,row_offset_in_EMU,padding_in_pixel=_DEFAULT_PADDING_IN_PIXEL):
        col_idx,row_idx = coordinate_string_to_index(cell_coord)
        marker = AnchorMarker(col=col_idx-1, colOff=col_offset_in_EMU+pixels_to_EMU(padding_in_pixel), row=row_idx-1, rowOff=row_offset_in_EMU+pixels_to_EMU(padding_in_pixel))
        size = XDRPositiveSize2D(pixels_to_EMU(image.width-padding_in_pixel), pixels_to_EMU(image.height-padding_in_pixel))
        image.anchor = OneCellAnchor(_from=marker,ext=size)
        sheet.add_image(image)
        # without padding:
        # sheet.add_image(image, target_cell_coord)

    @staticmethod
    def _points_to_pixels(points, conversion_factor=1.333):
        pixels = points * conversion_factor
        return pixels

    @staticmethod
    def _characters_to_pixels(character_width, font_size):
        # Convert character width to pixels
        pixels = character_width * _CAHR_WIDTH_PIXELS
        return pixels

    def _resize_image(self, image, desired_width=-1, desired_height=-1):
        original_width, original_height = image.width, image.height

        if desired_width == -1 and desired_height == -1:
            # Return the original image if no resizing is desired
            return image

        if desired_width == -1:
            new_width = int(desired_height *
                            (original_width / original_height))
            new_height = desired_height
        elif desired_height == -1:
            new_width = desired_width
            new_height = int(desired_width *
                             (original_height / original_width))
        else:
            # Resize with desired width and height
            new_width = desired_width
            new_height = desired_height

        image.height = new_height
        image.width = new_width

    def _resize_image_base_on_cell_size(self,image,cell_size_in_pixels):
        '''
        cell_size_in_pixels: a tuple (cell_width_in_pixels,cell_height_in_pixels)
        
        '''
        cell_width_in_pixels,cell_height_in_pixels = cell_size_in_pixels

        cell_aspect_ratio = cell_width_in_pixels/cell_height_in_pixels
        image_aspect_ratio = image.width/image.height
        if cell_aspect_ratio>image_aspect_ratio:
            self._resize_image(image,
                            desired_width=-1,
                            desired_height=cell_height_in_pixels)
        else:
            self._resize_image(image,
                            desired_width=cell_width_in_pixels,
                            desired_height=-1)

    def delete_temp_images():
        for temp_jpeg_path in self.temp_image_paths:
            os.remove(temp_jpeg_path)

if __name__ == '__main__':
    from openpyxl import Workbook, load_workbook
    template_workbook = load_workbook('/home/lighthouse/agenda_template_zoo/huangpu_rise_template_for_print/huangpu_rise_template_for_print.xlsx')
    template_sheet = template_workbook['page2']
    source_block_position = {'start_coord':'A1','end_coord':'L38'}
    target_coord = 'A1'
    writer = XlsxBlockImageWriter(template_sheet,source_block_position,target_coord)
    print(writer._get_image_source_coords())
    print(writer._get_image_coord_offset('B23'))
    print(writer._get_image_target_coords((11,0)))
