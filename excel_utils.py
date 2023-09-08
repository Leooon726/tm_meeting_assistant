from openpyxl.utils.cell import coordinate_from_string, get_column_letter, column_index_from_string


def coordinate_string_to_index(coord):
    coord_tuple = coordinate_from_string(
        coord)  # Parse the first coordinate string
    coord_col_idx = column_index_from_string(coord_tuple[0])
    return coord_col_idx, coord_tuple[1]


def index_to_coordinate_string(col_idx, row_idx):
    col_letter = get_column_letter(col_idx)
    return col_letter + str(row_idx)


def add_coordinates(coord, offsets):
    '''
    coord can be 'A1', offsets can be (1, 2)
    '''
    coord_col_idx, coord_row_idx = coordinate_string_to_index(coord)
    res_coord_col_idx = coord_col_idx + offsets[0]
    res_coord_row_idx = coord_row_idx + offsets[1]
    return index_to_coordinate_string(res_coord_col_idx, res_coord_row_idx)

def subtract_coordinates(coord1, coord2):
    coord1_col_idx, coord1_row_idx = coordinate_string_to_index(coord1)
    coord2_col_idx, coord2_row_idx = coordinate_string_to_index(coord2)
    return (coord1_col_idx - coord2_col_idx,
            coord1_row_idx - coord2_row_idx)

def get_left_top_coordinate(sheet):
    min_row = sheet.min_row
    min_col = sheet.min_column
    left_top_coordinate = f"{get_column_letter(min_col)}{min_row}"
    return left_top_coordinate


def get_right_bottom_coordinate(sheet):
    max_row = sheet.max_row
    max_col = sheet.max_column
    right_bottom_coordinate = f"{get_column_letter(max_col)}{max_row}"
    return right_bottom_coordinate


def get_sheet_dimensions(sheet):
    width = sheet.max_column - sheet.min_column + 1
    height = sheet.max_row - sheet.min_row + 1
    return width, height


def is_cell_in_block(src_start_coord, src_end_coord,
                        coord_to_check):
    min_col_idx, min_row_idx = coordinate_string_to_index(src_start_coord)
    max_col_idx, max_row_idx = coordinate_string_to_index(src_end_coord)
    cell_col_idx, cell_row_idx = coordinate_string_to_index(coord_to_check)
    return min_col_idx <= cell_col_idx <= max_col_idx and min_row_idx <= cell_row_idx <= max_row_idx


def calculate_block_size(start_coord,end_coord):
    '''
    start_coord and end_coord are like A10.
    '''
    start_col, start_row = coordinate_string_to_index(start_coord)
    end_col, end_row = coordinate_string_to_index(end_coord)
    return end_col-start_col+1,end_row-start_row+1


def get_merged_cell_size(worksheet, cell_coordinate):
    '''
    if the given cell coord belongs to a merged cell, return the merged cell's width and height.
    otherwise, return the given cell's width and height .
    '''
    for merged_range in worksheet.merged_cells.ranges:
        if cell_coordinate in merged_range:
            start_row, end_row = merged_range.min_row, merged_range.max_row
            start_col, end_col = merged_range.min_col, merged_range.max_col
            total_height = 0
            total_width = 0
            # Calculate total height
            for row in range(start_row, end_row + 1):
                total_height += worksheet.row_dimensions[row].height
            # Calculate total width
            for col in range(start_col, end_col + 1):
                column_letter = get_column_letter(col)
                total_width += worksheet.column_dimensions[column_letter].width
            return total_width,total_height
    cell = worksheet[cell_coordinate]
    cell_height = worksheet.row_dimensions[cell.row].height
    cell_width = worksheet.column_dimensions[cell.column_letter].width
    return cell_width,cell_height


if __name__ == '__main__':
    # assert calculate_block_size('A1','B1') == (2,1)

    import openpyxl
    workbook = openpyxl.load_workbook("/home/lighthouse/agenda_template_zoo/huangpu_rise_template_for_print/huangpu_rise_template_for_print.xlsx")
    sheet = workbook['page1']
    print(get_merged_cell_size(sheet,'A4'))