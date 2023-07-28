from openpyxl.cell import Cell
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, PatternFill
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl import Workbook, load_workbook
from copy import copy
from openpyxl.utils.cell import coordinate_from_string, get_column_letter,column_index_from_string,range_boundaries
from openpyxl.drawing.image import Image

def _coordinate_strig_to_index(coord):
    coord_tuple = coordinate_from_string(coord)  # Parse the first coordinate string
    coord_col_idx = column_index_from_string(coord_tuple[0])
    return coord_col_idx,coord_tuple[1]

def shift_coordinates(coord, offset):
    coord_col_idx,coord_row_idx = _coordinate_strig_to_index(coord)
    result_tuple = (coord_col_idx + offset[0], coord_row_idx + offset[1])  # Add the corresponding row and column values
    result_coord = get_column_letter(result_tuple[0]) + str(result_tuple[1])  # Convert the result back to a coordinate string
    return result_coord

def shift_range(range_string,offset):
    min_col, min_row, max_col, max_row = range_boundaries(range_string)
    result_tuple1 = (min_col + offset[0], min_row + offset[1])  # Add the corresponding row and column values
    result_coord1 = get_column_letter(result_tuple1[0]) + str(result_tuple1[1])  # Convert the result back to a coordinate string
    result_tuple2 = (max_col + offset[0], max_row + offset[1])  # Add the corresponding row and column values
    result_coord2 = get_column_letter(result_tuple2[0]) + str(result_tuple2[1])  # Convert the result back to a coordinate string
    return '{}:{}'.format(result_coord1,result_coord2)

def subtract_coordinates(coord1, coord2):
    coord1_col_idx,coord1_row_idx = _coordinate_strig_to_index(coord1)
    coord2_col_idx,coord2_row_idx = _coordinate_strig_to_index(coord2)
    return (coord1_col_idx-coord2_col_idx,coord1_row_idx-coord2_row_idx)

def xlsx_sheet_copy(src_path, src_sheet_name, tag_path, target_sheet_name,start_coord='A1'):  # 跨xlsx复制sheet
    # 跨xlsx文件复制源文件xlsx中指定的sheet
    # 保留所有格式，以及行高列宽，视觉效果几乎一致
    # 不能复制除了文字以外的东西，例如图片
    # src_path:源xlsx文件路径
    # tag_path:目标xlsx文件路径
    # sheet_name:需要复制的源xlsx文件sheet的名称
    src_workbook = load_workbook(src_path)  # 打开源xlsx
    src_file_sheet = src_workbook[src_sheet_name]  # 打开目标sheet
    try:
        tag_workbook = load_workbook(tag_path)  # Open the target workbook if it exists
    except FileNotFoundError:
        tag_workbook = Workbook()  # Create a new workbook if the target doesn't exist

    # Create or get the sheet with the desired sheet name
    if target_sheet_name in tag_workbook.sheetnames:
        tag_file_sheet = tag_workbook[target_sheet_name]
    else:
        tag_file_sheet = tag_workbook.create_sheet(target_sheet_name)

    shift_offset = subtract_coordinates(start_coord,'A1')
    for row in src_file_sheet:
        # 遍历源xlsx文件制定sheet中的所有单元格
        for cell in row:  # 复制数据
            target_cell_coord = shift_coordinates(cell.coordinate,shift_offset)
            tag_file_sheet[target_cell_coord].value = cell.value
            if cell.has_style:  # 复制样式
                tag_file_sheet[target_cell_coord].font = copy(cell.font)
                tag_file_sheet[target_cell_coord].border = copy(cell.border)
                tag_file_sheet[target_cell_coord].fill = copy(cell.fill)
                tag_file_sheet[target_cell_coord].number_format = copy(
                    cell.number_format
                )
                tag_file_sheet[target_cell_coord].protection = copy(cell.protection)
                tag_file_sheet[target_cell_coord].alignment = copy(cell.alignment)

    wm = list(zip(src_file_sheet.merged_cells))  # 开始处理合并单元格
    if len(wm) > 0:  # 检测源xlsx中合并的单元格
        for i in range(0, len(wm)):
            cell2 = (
                str(wm[i]).replace("(<MergedCellRange ", "").replace(">,)", "")
            )  # 获取合并单元格的范围
            shifted_cell2 = shift_range(cell2,shift_offset)
            tag_file_sheet.merge_cells(shifted_cell2)  # 合并单元格
    # 开始处理行高列宽
    col_shift_offset, row_shift_offset = shift_offset
    for i in range(1, src_file_sheet.max_row + 1):
        tag_file_sheet.row_dimensions[i+row_shift_offset].height = src_file_sheet.row_dimensions[
            i
        ].height
    for i in range(1, src_file_sheet.max_column + 1):
        tag_file_sheet.column_dimensions[
            get_column_letter(i+col_shift_offset)
        ].width = src_file_sheet.column_dimensions[get_column_letter(i)].width

    tag_workbook.save(tag_path)  # 保存
    tag_workbook.close()  # 关闭文件
    src_workbook.close()


def add_image_to_excel(file_path, sheet_name, image_path, target_cell_coord):
    # Load the workbook
    workbook = load_workbook(file_path)

    try:
        # Get the sheet
        sheet = workbook[sheet_name]

        # Load the image
        image = Image(image_path)

        # col_idx,row_idx = _coordinate_strig_to_index(target_cell_coord)
        # col_letter = get_column_letter(col_idx)
        # cell_width = sheet.column_dimensions[col_letter].width
        # cell_height = sheet.row_dimensions[row_idx].height
        image.width = 100
        image.height = 100

        # Add the image to the worksheet
        sheet.add_image(image, target_cell_coord)

        # Save the workbook
        workbook.save(file_path)

        print(f"Image added to '{sheet_name}' in {file_path}")

    except KeyError:
        print(f"Sheet '{sheet_name}' does not exist in the workbook.")

    # Close the workbook
    workbook.close()


def find_text_in_cells(file_path, sheet_name, target_text):
    # Load the workbook and select the desired worksheet
    workbook = load_workbook(file_path)
    worksheet = workbook[sheet_name]

    # List to store the coordinates of cells with the target text
    cell_coordinates = []

    max_row = worksheet.max_row
    max_column = worksheet.max_column

    for row_num in range(1, max_row + 1):
        for col_num in range(1, max_column + 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell_value = cell.value
            if isinstance(cell_value, str) and target_text in cell_value:
                cell_coordinates.append(cell.coordinate)

    return cell_coordinates


target_sheet_name = 'calendar'

source_file = '/home/didi/myproject/tmma/tm_303_calendar.xlsx'
target_file = '/home/didi/myproject/tmma/generated_calendar.xlsx'
title_block = 'title_block'
theme_block = 'theme_block'
parent_event = 'parent_event'
child_event = 'child_event'
notice_block = 'notice_block'
rule_block = 'rule_block'
information_block = 'information_block'

# xlsx_sheet_copy(source_file,sheet1_name,target_file,target_sheet_name)
# xlsx_sheet_copy(source_file,sheet2_name,target_file,target_sheet_name,start_coord='B12')
# xlsx_sheet_copy(source_file,sheet3_name,target_file,target_sheet_name,start_coord='L9')
# xlsx_sheet_copy(source_file,sheet4_name,target_file,target_sheet_name,start_coord='L16')
# image_path = '/home/didi/myproject/tmma/tm_logo.jpg'
# add_image_to_excel(target_file,target_sheet_name,image_path,'B2')
# res = find_text_in_cells(source_file,sheet2_name,'{event_name}')
###########test
# xlsx_sheet_copy(source_file,target_file,sheet2_name)
# print(shift_range('A34:F34',(1,2)))
