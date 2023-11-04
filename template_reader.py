import re

from openpyxl import load_workbook

from excel_utils import calculate_block_size,coordinate_string_to_index, add_coordinates, get_left_top_coordinate, get_sheet_dimensions, get_right_bottom_coordinate


class XlsxTemplateReader():

    def __init__(self, template_xlsx_file_path, template_sheet_name,
                 position_sheet_name):
        self.template_workbook = load_workbook(template_xlsx_file_path)
        self.template_sheet = self.template_workbook[template_sheet_name]
        self.template_position_sheet = self.template_workbook[
            position_sheet_name]
        self.template_block_position_dict = self._calculate_block_position()
        self.template_block_size_dict = self._calcuate_block_sizes()

    def __exit__(self):
        self.template_workbook.close()

    @staticmethod
    def is_legal_coord(coord):
        if not isinstance(coord,str) or coord is None:
            return False
        if ',' in coord:
            return False
        pattern = r'^[A-Z]{1,2}[0-9]+$'
        return bool(re.match(pattern, coord))

    def _calcuate_block_sizes(self):
        size_dict = {}
        for block_name,block_position  in self.template_block_position_dict.items():
            if not self.is_legal_coord(block_position['start_coord']) or not self.is_legal_coord(block_position['end_coord']):
                continue
            w,h = calculate_block_size(block_position['start_coord'],block_position['end_coord'])
            size_dict[block_name] = dict(width=w,height=h)
        return size_dict

    @staticmethod
    def _extract_field(input_string):
        '''
        input_string: like "会议主题：{主题%父亲节}" or "会议主题：{主题}"
        return: "主题"
        '''
        start = input_string.find('{')
        end = input_string.find('%') if '%' in input_string else input_string.find('}')

        if start != -1 and end != -1 and start < end:
            # Do not containt the '{', '}' and '%'.
            extracted_text = input_string[start+1:end]
            return extracted_text
        else:
            return None

    @staticmethod
    def _extract_field_with_example(input_string):
        '''
        Example 1:
            input_string: "会议主题：{主题}"
            return: {'field_name':"主题"}
        Example 2:
            input_string: "会议主题：{主题%父亲节}"
            return: {'field_name':"主题", 'example':"父亲节"}
        '''
        field_name = XlsxTemplateReader._extract_field(input_string)
        if field_name is None:
            return None

        res = {'field_name':field_name}
        if '%' in input_string:
            start = input_string.find('%')
            end = input_string.find('}')
            example = input_string[start+1:end]
            res['example'] = example
        res['is_single_line'] = ('\n' not in input_string)
        return res

    def get_user_filled_fields_from_sheet(self):
        field_list = []
        for block_name in self.template_block_position_dict.keys():
            # the following blocks should be automatically filled by program, not by users.
            if block_name in ['parent_block','notice_block','child_block']:
                continue
            field_list+=self.get_field_list(block_name,extract_method=XlsxTemplateReader._extract_field_with_example)
        return field_list

    def get_field_list(self, template_block_name,extract_method=None):
        position = self.template_block_position_dict[template_block_name]
        if position is None:
            start_coord = get_left_top_coordinate(self.template_sheet)
            end_coord = get_right_bottom_coordinate(self.template_sheet)
        else:
            start_coord = position['start_coord']
            end_coord = position['end_coord']

        start_col, start_row = coordinate_string_to_index(start_coord)
        end_col, end_row = coordinate_string_to_index(end_coord)
        field_list = []
        if extract_method is None:
            extract_method = XlsxTemplateReader._extract_field
        for row_num in range(start_row, end_row + 1):
            for col_num in range(start_col, end_col + 1):
                cell = self.template_sheet.cell(row=row_num, column=col_num)
                if cell.value is not None and isinstance(
                        cell.value, str) and '{' in cell.value:
                    field_list.append(extract_method(cell.value))
        return field_list

    def _read_sheet_as_dict_list(self):
        headers = [cell.value for cell in self.template_position_sheet[1]]

        column_as_key_list = []
        for row in self.template_position_sheet.iter_rows(min_row=2,
                                                          values_only=True):
            row_data = {}
            for header, cell_value in zip(headers, row):
                row_data[header] = cell_value
            column_as_key_list.append(row_data)
        return column_as_key_list

    def _calculate_block_position(self):
        '''
        return a double-layered dict, like:
        {'title_block': {'start_coord': 'A1', 'end_coord': 'N7'}, 
        'theme_block': {'start_coord': 'A8', 'end_coord': 'J10'}}
        '''
        column_as_key_list = self._read_sheet_as_dict_list()
        result_dict = {}
        for item in column_as_key_list:
            result_dict[item['block_name']] = {
                'start_coord': item['start_coord'],
                'end_coord': item['end_coord']
            }
        return result_dict

    def get_template_block_sizes(self):
        return self.template_block_size_dict

    def get_template_positions(self):
        return self.template_block_position_dict

    def is_pure_text_block(self, template_block_name):
        return len(self.get_field_list(template_block_name)) == 0
    
    def get_image_coords(self):
        column_as_key_list = self._read_sheet_as_dict_list()
        for row_dict in column_as_key_list:
            if row_dict['block_name'] == 'images':
                return row_dict['start_coord'].split(',')

if __name__ == '__main__':
    template_file = '/home/lighthouse/tm_meeting_assistant/example/jabil_jouse_template_for_print/jabil_jouse_template_for_print.xlsx'
    template_sheet_name = 'template'
    template_position_sheet_name = 'template_position'
    # template_file = '/home/lighthouse/agenda_template_zoo/huangpu_rise_template_for_print/huangpu_rise_template_for_print.xlsx'
    # template_sheet_name = 'page1'
    # template_position_sheet_name = 'page1_block_position'

    tr = XlsxTemplateReader(template_file, template_sheet_name,
                            template_position_sheet_name)
    print(tr.get_user_filled_fields_from_sheet())
    print(tr.template_block_size_dict)

    field_list = tr.get_field_list('contact_block')
    assert field_list == []
    
    field_list = tr.get_field_list('title_block')
    print(field_list)
    print(tr.get_template_positions())
    print(tr.is_pure_text_block('contact_block'))
    print(tr.get_image_coords())
