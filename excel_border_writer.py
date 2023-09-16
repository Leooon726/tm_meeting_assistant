import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Border, Side

def set_outer_borders(excel_file_path='/home/lighthouse/output_files/adhoc/jabil_test.xlsx',sheet_name='Sheet',output_file_path=None):
    # Load the workbook
    workbook = load_workbook(excel_file_path)

    # Choose the worksheet
    worksheet = workbook[sheet_name]


    # Get the dimensions of the worksheet
    dimensions = worksheet.dimensions
    min_column = worksheet.min_column
    max_column = worksheet.max_column
    min_row = worksheet.min_row
    max_row = worksheet.max_row


    right_border_style = Border(
        left=None,
        right=Side(border_style='thin', color='000000'),
        top=None,
        bottom=None
    )
    left_border_style = Border(
        left=Side(border_style='thin', color='000000'),
        right=None,
        top=None,
        bottom=None
    )
    for row in range(min_row, max_row + 1):
        # Add right border for the right-edge cells.
        right_cell = worksheet.cell(row=row, column=max_column)
        if right_cell.border.right.border_style is None:
            right_cell.border = right_border_style
        # Add left border for the left-edge cells.
        left_cell = worksheet.cell(row=row, column=min_column)
        if left_cell.border.left.border_style is None:
            left_cell.border = left_border_style

    top_border_style = Border(
        left=None,
        right=None,
        top=Side(border_style='thin', color='000000'),
        bottom=None
    )
    bottom_border_style = Border(
        left=None,
        right=None,
        top=None,
        bottom=Side(border_style='thin', color='000000')
    )
    for column in range(min_column, max_column + 1):
        # Add top border for the top-edge cells.
        top_cell = worksheet.cell(row=min_row, column=column)
        if top_cell.border.top.border_style is None:
            top_cell.border = top_border_style

        # Add bottom border for the bottom-edge cells.
        bottom_cell = worksheet.cell(row=max_row, column=column)
        if bottom_cell.border.bottom.border_style is None:
            bottom_cell.border = bottom_border_style

    # Save the workbook
    workbook.save(output_file_path)

if __name__ == '__main__':
    set_outer_borders(output_file_path='/home/lighthouse/output_files/adhoc/jabil_border_test.xlsx')
