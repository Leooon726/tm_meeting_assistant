import os
import uuid
from os.path import isfile
import tempfile
import shutil

from openpyxl.cell import Cell
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, PatternFill
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl import Workbook, load_workbook
from copy import copy, deepcopy
from openpyxl.utils.cell import coordinate_from_string, get_column_letter, column_index_from_string, range_boundaries
from openpyxl.drawing.image import Image

from excel_utils import get_merged_cell_size,coordinate_string_to_index, add_coordinates, get_left_top_coordinate, get_sheet_dimensions, get_right_bottom_coordinate,is_cell_in_block,subtract_coordinates
from xlsx_image_writer import XlsxBlockImageWriter
from sheet_image_loader import SheetImageLoader

class XlsxWriter():

    def __init__(self, template_xlsx_file_path, template_sheet_name,
                 template_position_config, target_xlsx_file_path,
                 target_sheet_name):
        self.template_workbook = load_workbook(template_xlsx_file_path)
        self.template_sheet = self.template_workbook[template_sheet_name]
        self.template_position_config = template_position_config

        if isfile(target_xlsx_file_path):
            # If the file exists, load the existing workbook
            self.target_workbook = load_workbook(target_xlsx_file_path)
        else:
            # If the file doesn't exist, create a new workbook
            self.target_workbook = Workbook()
            # And delete the default sheet.
            default_sheet = self.target_workbook.active
            self.target_workbook.remove(default_sheet)

        self.target_xlsx_file_path = target_xlsx_file_path
        self.temp_image_paths = []
        # Create or get the sheet with the desired sheet name
        if target_sheet_name in self.target_workbook.sheetnames:
            self.target_sheet = self.target_workbook[target_sheet_name]
        else:
            self.target_sheet = self.target_workbook.create_sheet(
                target_sheet_name)

        self.temp_dir = tempfile.mkdtemp()

        self.template_image_loader = SheetImageLoader(self.template_sheet,self.temp_dir)

    def save(self):
        self.target_workbook.save(self.target_xlsx_file_path)  # 保存
        self.target_workbook.close()  # 关闭文件
        self.template_workbook.close()

        if self.temp_dir:
            # Remove the temporary directory and all its contents
            shutil.rmtree(self.temp_dir)
            self.temp_dir = None

    def write_sheet(self,
                    source_template_block_name,
                    target_start_coord='A1',
                    data=None):
        source_position = self.template_position_config[
            source_template_block_name]
        src_file_sheet = self._copy_sheet_impl(self.template_sheet,
                                               source_position)
        if data is not None:
            src_file_sheet = self._modify_sheet(src_file_sheet, data)
        self._paste_sheet_impl(self.target_sheet, target_start_coord,
                               src_file_sheet, source_position)

        # write images.
        image_writer = XlsxBlockImageWriter(self.template_sheet,source_position,target_start_coord,self.template_image_loader)
        self.target_sheet = image_writer.write(self.target_sheet)

    def _shift_coordinates(self, coord, offset):
        coord_col_idx, coord_row_idx = coordinate_string_to_index(coord)
        result_tuple = (coord_col_idx + offset[0], coord_row_idx + offset[1]
                        )  # Add the corresponding row and column values
        result_coord = get_column_letter(result_tuple[0]) + str(
            result_tuple[1])  # Convert the result back to a coordinate string
        return result_coord

    def _shift_range(self, range_string, offset):
        min_col, min_row, max_col, max_row = range_boundaries(range_string)
        result_tuple1 = (min_col + offset[0], min_row + offset[1]
                         )  # Add the corresponding row and column values
        result_coord1 = get_column_letter(result_tuple1[0]) + str(
            result_tuple1[1])  # Convert the result back to a coordinate string
        result_tuple2 = (max_col + offset[0], max_row + offset[1]
                         )  # Add the corresponding row and column values
        result_coord2 = get_column_letter(result_tuple2[0]) + str(
            result_tuple2[1])  # Convert the result back to a coordinate string
        return '{}:{}'.format(result_coord1, result_coord2)

    def _copy_sheet_impl(self, original_sheet, source_position=None):
        if source_position is None:
            start_coord = get_left_top_coordinate(original_sheet)
            end_coord = get_right_bottom_coordinate(original_sheet)
        else:
            start_coord = source_position['start_coord']
            end_coord = source_position['end_coord']

        new_workbook = Workbook()
        new_workbook.remove(new_workbook.active)  # Remove the default sheet

        new_sheet = new_workbook.create_sheet(original_sheet.title)
        start_col, start_row = coordinate_string_to_index(start_coord)
        end_col, end_row = coordinate_string_to_index(end_coord)
        for row in original_sheet.iter_rows(min_row=start_row,
                                            min_col=start_col,
                                            max_row=end_row,
                                            max_col=end_col):
            # 遍历源xlsx文件制定sheet中的所有单元格
            for cell in row:  # 复制数据
                target_cell_coord = cell.coordinate
                new_sheet[target_cell_coord].value = cell.value
                if cell.has_style:  # 复制样式
                    new_sheet[target_cell_coord].font = copy(cell.font)
                    new_sheet[target_cell_coord].border = copy(cell.border)
                    new_sheet[target_cell_coord].fill = copy(cell.fill)
                    new_sheet[target_cell_coord].number_format = copy(
                        cell.number_format)
                    new_sheet[target_cell_coord].protection = copy(
                        cell.protection)
                    new_sheet[target_cell_coord].alignment = copy(
                        cell.alignment)

        wm = list(zip(original_sheet.merged_cells))  # 开始处理合并单元格
        if len(wm) > 0:  # 检测源xlsx中合并的单元格
            for i in range(0, len(wm)):
                cell2 = (str(wm[i]).replace("(<MergedCellRange ",
                                            "").replace(">,)",
                                                        ""))  # 获取合并单元格的范围
                shifted_cell2 = cell2
                new_sheet.merge_cells(shifted_cell2)  # 合并单元格
        # 开始处理行高列宽
        for i in range(1, original_sheet.max_row + 1):
            new_sheet.row_dimensions[i].height = original_sheet.row_dimensions[
                i].height
        for i in range(1, original_sheet.max_column + 1):
            new_sheet.column_dimensions[get_column_letter(
                i)].width = original_sheet.column_dimensions[get_column_letter(
                    i)].width

        return new_sheet

    def _paste_sheet_impl(self,
                          target_file_sheet,
                          target_start_coord,
                          src_file_sheet,
                          source_position=None):
        if source_position is None:
            src_start_coord = get_left_top_coordinate(src_file_sheet)
            src_end_coord = get_right_bottom_coordinate(src_file_sheet)
        else:
            src_start_coord = source_position['start_coord']
            src_end_coord = source_position['end_coord']

        shift_offset = subtract_coordinates(target_start_coord,
                                                  src_start_coord)
        src_start_col, src_start_row = coordinate_string_to_index(
            src_start_coord)
        src_end_col, src_end_row = coordinate_string_to_index(src_end_coord)
        for row in src_file_sheet.iter_rows(min_row=src_start_row,
                                            min_col=src_start_col,
                                            max_row=src_end_row,
                                            max_col=src_end_col):
            # 遍历源xlsx文件制定sheet中的所有单元格
            for cell in row:  # 复制数据
                target_cell_coord = self._shift_coordinates(
                    cell.coordinate, shift_offset)
                # It would fail to write if the target cell is a merged cell which is read-only.
                try:
                    target_file_sheet[target_cell_coord].value = cell.value
                except:
                    # TODO: only capture read-only-error and print it.
                    pass
                    # print('Fail to write data for coord: ',target_cell_coord, ' with data: ',cell.value)
                if cell.has_style:  # 复制样式
                    target_file_sheet[target_cell_coord].font = copy(cell.font)
                    target_file_sheet[target_cell_coord].border = copy(
                        cell.border)
                    target_file_sheet[target_cell_coord].fill = copy(cell.fill)
                    target_file_sheet[target_cell_coord].number_format = copy(
                        cell.number_format)
                    target_file_sheet[target_cell_coord].protection = copy(
                        cell.protection)
                    target_file_sheet[target_cell_coord].alignment = copy(
                        cell.alignment)

        wm = list(zip(src_file_sheet.merged_cells))  # 开始处理合并单元格
        if len(wm) > 0:  # 检测源xlsx中合并的单元格
            for i in range(0, len(wm)):
                cell_range = (str(wm[i]).replace("(<MergedCellRange ",
                                                 "").replace(">,)",
                                                             ""))  # 获取合并单元格的范围
                begin_coord, end_coord = cell_range.split(':')
                if not is_cell_in_block(
                        src_start_coord, src_end_coord,
                        begin_coord) and not is_cell_in_block(
                            src_start_coord, src_end_coord, end_coord):
                    continue
                shifted_cell_range = self._shift_range(cell_range,
                                                       shift_offset)
                target_file_sheet.merge_cells(shifted_cell_range)  # 合并单元格
        # 开始处理行高列宽
        col_shift_offset, row_shift_offset = shift_offset
        for i in range(src_start_row, src_end_row + 1):
            target_file_sheet.row_dimensions[
                i + row_shift_offset].height = src_file_sheet.row_dimensions[
                    i].height
        for i in range(src_start_col, src_end_col + 1):
            target_file_sheet.column_dimensions[get_column_letter(
                i +
                col_shift_offset)].width = src_file_sheet.column_dimensions[
                    get_column_letter(i)].width

    @staticmethod
    def _is_field_to_be_filled(cell_string,candidate_key):
        '''
        cell_string: like "主题：{主题%父亲节}" or "主题：{主题}"
        if |candidate_key| is like "主题", then returns True.
        otherwise returns False.
        '''
        if '{' not in cell_string or '}' not in cell_string:
            return False
        # Extract the field name from the cell string
        field_name_start = cell_string.find('{') + 1
        field_name_end = cell_string.find('%') if '%' in cell_string else cell_string.find('}')
        field_name = cell_string[field_name_start:field_name_end]

        # Compare the extracted field name with the candidate key
        if field_name == candidate_key:
            return True
        else:
            return False

    @staticmethod
    def _fill_field(cell_string,candidate_value):
        '''
        cell_string: like "主题：{主题%父亲节}" or "主题：{主题}"
        candidate_value: like "旅行"
        return: "主题：旅行"
        String within {} is replaced.
        '''
        candidate_value = str(candidate_value)

        if '{' not in cell_string or '}' not in cell_string:
            return cell_string

        start = cell_string.find('{')
        end = cell_string.find('}')
        string_to_be_replaced = cell_string[start:end+1]

        return cell_string.replace(string_to_be_replaced, candidate_value)

    def _modify_sheet(self, sheet, data):
        max_row = sheet.max_row
        max_column = sheet.max_column

        for key, value in data.items():
            for row_num in range(1, max_row + 1):
                for col_num in range(1, max_column + 1):
                    cell = sheet.cell(row=row_num, column=col_num)
                    cell_value = cell.value
                    if cell_value is not None and isinstance(cell_value,str) and self._is_field_to_be_filled(cell_value,key):
                        cell.value = self._fill_field(cell_value,value)
                        # cell.value = cell.value.replace(key, str(value))
        return sheet

    def merge_cells(self, start_coord, end_coord):
        cell_range = start_coord + ':' + end_coord
        self.target_sheet.merge_cells(cell_range)


if __name__ == '__main__':
    # tests
    # res = add_coordinates('B12',(0,1))

    template_file = '/home/didi/myproject/tmma/tm_303_calendar.xlsx'
    template_sheet_name = 'template'
    target_file = '/home/didi/myproject/tmma/generated_calendar.xlsx'
    target_sheet_name = 'Sheet'
    image_path = '/home/didi/myproject/tmma/tm_logo.jpg'

    title_block = 'title_block'
    theme_block = 'theme_block'
    parent_block = 'parent_block'
    child_block = 'child_block'
    notice_block = 'notice_block'
    rule_block = 'rule_block'
    information_block = 'information_block'

    template_position_config = {
        'title_block': {
            'start_coord': 'B2',
            'end_coord': 'O8'
        },
        'theme_block': {
            'start_coord': 'B9',
            'end_coord': 'K11'
        },
        'parent_block': {
            'start_coord': 'B12',
            'end_coord': 'K12'
        },
        'child_block': {
            'start_coord': 'B13',
            'end_coord': 'K13'
        }
    }

    xlsx_writer = XlsxWriter(template_file, template_sheet_name,
                             template_position_config, target_file,
                             target_sheet_name)
    xlsx_writer.write_sheet(source_template_block_name=title_block,
                            target_start_coord='B2')
    xlsx_writer.write_sheet(source_template_block_name=theme_block,
                            target_start_coord='B9',
                            data={
                                '{theme_name}': '主题：父亲节',
                                '{time}': '15:00',
                                '{organizer_name}': '林长虹',
                                '{SAA_name}': 'Leon Lin'
                            })
    cur_start_coord = 'B12'
    for i in range(3):
        cur_start_coord = add_coordinates(cur_start_coord, (0, 1))
        xlsx_writer.write_sheet(source_template_block_name=parent_block,
                                target_start_coord=cur_start_coord,
                                data={
                                    '{start_time}': '12:01',
                                    '{event_name}': '开场',
                                    '{duration}': '8'
                                })
        cur_start_coord = add_coordinates(cur_start_coord, (0, 1))
        xlsx_writer.write_sheet(source_template_block_name=child_block,
                                target_start_coord=cur_start_coord,
                                data={
                                    '{event_name}': '开场白',
                                    '{end_time}': '15:01',
                                    '{duration}': '8',
                                    '{host_name}': '林长宏'
                                })
    # xlsx_writer.write_sheet('calendar',source_position=template_position_config[rule_block],target_start_coord='L9')
    # xlsx_writer.write_sheet('calendar',source_position=template_position_config[information_block],target_start_coord='L16')
    xlsx_writer.add_image(source_cell_coord='B2',
                          target_cell_coord='B2',
                          image_width=150)
    xlsx_writer.add_image(source_cell_coord='O2',
                          target_cell_coord='O2',
                          image_width=150)
    xlsx_writer.save()
